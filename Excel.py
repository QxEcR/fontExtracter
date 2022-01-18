import openpyxl
from pathlib import Path

class Excel:
    def __init__(self, filename):
        self.filename = filename
        xlsx_path = Path(self.filename)
        self.xlsx_file = openpyxl.load_workbook(xlsx_path) 

        self.inputSheet = self.xlsx_file['Input']
        self.outputSheet = self.xlsx_file['Output']
    
    def getInputs(self):
        list = []
        for row in self.inputSheet.iter_rows(max_row=self.inputSheet.max_row):
            for cell in row:
                if cell.value.startswith("http"):
                    list.append(cell.value)
        return list

    def writeToExcel(self, url, isCSS, list):
        for item in list:
            self.outputSheet.append((url, "CSS" if isCSS else "Non-CSS", item))
        self.xlsx_file.save(self.filename)